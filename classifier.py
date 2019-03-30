#KFold
from sklearn.model_selection import KFold
#classifiers
from sklearn.tree import DecisionTreeClassifier
from sklearn import svm
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import RandomForestClassifier,BaggingClassifier,AdaBoostClassifier
from sklearn.neighbors import KNeighborsClassifier
#sampling algorithms
from imblearn.over_sampling import SMOTE,ADASYN
from imblearn.combine import SMOTEENN,SMOTETomek
#miscellaneous imports
from openpyxl import Workbook
import os
import csv
import numpy
classifier_list=['Decision Tree','Naive Bayes','KNN','Random Forest','Bagging','AdaBoost']

# dataset , sampling_index,clf name 

def write_results_in_workbook(result_list):
	wb = Workbook()
	ws = wb.active
	ws.cell(row=1,column=1).value="Datasets"
	ws.merge_cells('B1:G1')
	ws.merge_cells('H1:M1')
	ws.merge_cells('N1:S1')
	ws.merge_cells('T1:Y1')
	ws.merge_cells('Z1:AE1')
	ws.cell(row=1,column=2).value="NONE"
	ws.cell(row=1,column=8).value="SMOTE"
	ws.cell(row=1,column=14).value="SMOTEENN"
	ws.cell(row=1,column=20).value="ADASYN"
	ws.cell(row=1,column=26).value="TOMEK"
	index = 2 
	for i in range(5):
		for clf in classifier_list:
			ws.cell(row=2,column=index).value=clf
			index +=1

	for i in range(len(result_list)):
		ws.cell(row=i+3,column=1).value=datasets[i]
		index = 2 
		for it in range(5):
			for clf in classifier_list:
				ws.cell(row=i+3,column=index).value=result_list[i][it][sampling_algorithm_list[it]][clf]
				index +=1
	wb.save('newresult.xlsx')


def load_Data(dataset):
	data=[]	
	target=[]
	dataset = "dataset/" + dataset+ ".csv"
	with open(dataset,"r") as csvfile:
		csvreader = csv.reader(csvfile)
		fields = csvreader.next()
		for row in csvreader:
			lst=[]
			for i in range(len(row)-1):
				lst.append(eval(row[i]))
			data.append(lst)
			target.append(row[len(row)-1])
	return data,target


def removenoise(data,target):
	gnb = GaussianNB()
	gnb.fit(data,target)
	predicted = gnb.predict(data)
	ln = len(target)
	filtered_data=[]
	filtered_target=[]
	for i in range(ln):
		if target[i] == predicted[i]:
			filtered_data.append(data[i])
			filtered_target.append(target[i])
	#print(len(target),len(filtered_target))
	return filtered_data,filtered_target


sampling_algorithm_list=["NONE","SMOTE","SMOTEENN","ADASYN","TOMEK"]


def check_data(train,target):
	index={}
	for ind in range(len(target)):
		if index.get(target[ind])==None:
			index[target[ind]] = ind 
	for x,y in index.items():
		for i in range(10):
			train.append(train[y])
			target.append(target[y])
	return train,target

def apply_sampling(train,target,index):
	train,target=check_data(train,target)
	if index==0:
		return train,target
	elif index==1:
		return SMOTE().fit_resample(numpy.asarray(train),numpy.asarray(target))
	elif index==2:
		return SMOTEENN().fit_resample(numpy.asarray(train),numpy.asarray(target))
	elif index==3:
		return ADASYN().fit_resample(numpy.asarray(train),numpy.asarray(target))
	else:
		return SMOTETomek().fit_resample(numpy.asarray(train),numpy.asarray(target))



def get_class_info(name,target):
	m={}
	for i in target:
		if m.get(i)==None:
			m[i]=1
		else:
			m[i] +=1
	print(name,m)

class Models:
	def __init__(self,name,index):
		self.name = name
		self.index = index
		self.DTClassifier = DecisionTreeClassifier()
		self.NBClassifier = GaussianNB()
		self.SVMClassifier = svm.SVC()
		self.KNNClassifier = KNeighborsClassifier()
		self.RFClassifier = RandomForestClassifier()
		self.BGClassifier = BaggingClassifier()
		self.ADBClassifier = AdaBoostClassifier()
		self.accuracy={}
		self.accuracy['Decision Tree'] = 0.0
		self.accuracy['Naive Bayes'] = 0.0
		self.accuracy['KNN']=0.0
		#self.accuracy['SVM'] = 0.0
		self.accuracy['Random Forest'] = 0.0
		self.accuracy['Bagging'] = 0.0
		self.accuracy['AdaBoost'] = 0.0

	def fit(self,data,target):
		self.DTClassifier.fit(data,target)
		self.NBClassifier.fit(data,target)
		self.KNNClassifier.fit(data,target)
		#self.SVMClassifier.fit(data,target)
		self.RFClassifier.fit(data,target)
		self.BGClassifier.fit(data,target)
		self.ADBClassifier.fit(data,target)

	def fit_score(self,train,target_train,test,target_test):
		train,target_train = removenoise(train,target_train)
		train,target_train = apply_sampling(train,target_train,self.index)
		#get_class_info(self.name,target_train)
		self.fit(train,target_train)
		self.accuracy['Decision Tree'] += self.DTClassifier.score(test,target_test)
		self.accuracy['Naive Bayes'] += self.NBClassifier.score(test,target_test)
		#self.accuracy['SVM'] += self.SVMClassifier.score(test,target_test)
		self.accuracy['KNN'] += self.KNNClassifier.score(test,target_test)
		self.accuracy['Random Forest'] += self.RFClassifier.score(test,target_test)
		self.accuracy['Bagging'] += self.BGClassifier.score(test,target_test)
		self.accuracy['AdaBoost'] += self.ADBClassifier.score(test,target_test)

	def get_accuracy(self):
		self.accuracy['Decision Tree'] *= 10 
		self.accuracy['Naive Bayes'] *= 10.0 
		#self.accuracy['SVM'] *=10.0 
		self.accuracy['KNN'] *= 10.0
		self.accuracy['Random Forest'] *= 10.0 
		self.accuracy['Bagging'] *= 10.0 
		self.accuracy['AdaBoost'] *= 10.0 
		m={}
		m[self.name]=self.accuracy
		return m



def applyKFold(data,target,remove_noise):
	acc=0	
	kf = KFold(n_splits=10,shuffle=True)
	classifier_models_none = Models('NONE',0)
	classifier_models_smote = Models('SMOTE',1)
	classifier_models_smoteenn = Models('SMOTEENN',2)
	classifier_models_adasyn = Models('ADASYN',3)
	classifier_models_tomek = Models('TOMEK',4)
	i=0
	for train_indices,test_indices in kf.split(data):
		train_data=[]
		train_target=[]
		test_data=[]
		test_target=[]
		i+=1
		for index in train_indices:
			train_data.append(data[index])
			train_target.append(target[index])
		
		for index in test_indices:
			test_data.append(data[index])
			test_target.append(target[index])
		
		# removing noise in the below code using naive_bayes classifier
		"""if remove_noise == True:
			train_data,train_target=removenoise(train_data,train_target)
		"""
		#training the model
		classifier_models_none.fit_score(train_data,train_target,test_data,test_target)
		classifier_models_smote.fit_score(train_data,train_target,test_data,test_target)
		classifier_models_smoteenn.fit_score(train_data,train_target,test_data,test_target)
		classifier_models_adasyn.fit_score(train_data,train_target,test_data,test_target)
		classifier_models_tomek.fit_score(train_data,train_target,test_data,test_target)
	accuracy_list = [classifier_models_none.get_accuracy(),classifier_models_smote.get_accuracy(),classifier_models_smoteenn.get_accuracy(),
	classifier_models_adasyn.get_accuracy(),classifier_models_tomek.get_accuracy()]
	return accuracy_list


#The program starts here
datasets=["haberman","yeast","cleveland","pima","thyroid","segment"]#["haberman","bupa","breast","cleveland","pima","thyroid"]
result_list=[]
for dataset in datasets:
	print(dataset)
	data,target = load_Data(dataset)
	acc=applyKFold(data,target,remove_noise=True)
	for row in acc:
		print(row)
	result_list.append(acc)
write_results_in_workbook(result_list)
#haberman,yeast,cleveland,pima,thyroid,segment